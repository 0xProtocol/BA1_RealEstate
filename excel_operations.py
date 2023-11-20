# excel_operations.py
from openpyxl import load_workbook

def load_excel_data(file_path='data.xlsx'):
    wb = load_workbook(file_path)
    ws = wb.active
    return ws

def extract_column_values(ws, column_index, start_row=2):
    max_row = ws.max_row
    column_values = []

    for row in range(start_row, max_row):
        cell_value = ws.cell(row=row, column=column_index).value
        column_values.append(cell_value)

    return column_values
