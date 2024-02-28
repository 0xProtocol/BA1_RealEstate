from openpyxl import load_workbook


def load_excel_data(file_path):
    # load the excel workbook
    wb = load_workbook(file_path)
    # select the active worksheet
    ws = wb.active
    # determine the maximum number of rows in the worksheet
    max_row = ws.max_row
    # create an empty list to store all rows from the excel file
    all_rows = []

    # iterate over the rows, starting from row 2 to the maximum row
    for row in range(2, max_row):
        # Create an array for the current row
        current_row = []

        # iterate over columns from B to Y (column index 1 to 25)
        for col in range(1, 26):  # 26 corresponds to the column index for Y
            cell_value = ws.cell(row=row, column=col).value
            current_row.append(cell_value)

        # append the current row to the list of all rows
        all_rows.append(current_row)

    return all_rows


if __name__ == "__main__":
    file_path = 'data.xlsx'
    data = load_excel_data(file_path)

    for row in data:
        print(row)
