from openpyxl import Workbook, load_workbook

wb = load_workbook('data.xlsx')
ws = wb.active

class Color:
    RESET = '\033[0m'
    RED = '\033[91m'
    GREEN = '\033[92m'

#import (Fragbg. Nr) ID?

#import (Tatsä. Nr) for ID
max_row = ws.max_row
all_rows = []

a_values = [] #Fragebogennummer
b_values = [] #Tatsä. Nr.
c_values = [] #Datum
d_values = [] #Ort
e_values = [] #Befrager
f_values = [] #Wohnung Zufr.
g_values = [] #Art Wohnung
h_values = [] #Miethöhe
i_values = [] #Wohnungssuche Optimismus
j_values = [] #Wohnpolitik
k_values = [] #Zur Verfügung
l_values = [] #Umgang Verm.
m_values = [] #Verm. Arbeit
n_values = [] # ??????????
o_values = [] #Makler harte Arb.
p_values = [] # ??????????
q_values = [] #Mak. wichtige Arb.
r_values = [] # ??????????
s_values = [] #Vermieter Beruf
t_values = [] #Einstl. Verm.
u_values = [] #Mietendeckel
v_values = [] #Einstl. Makler
w_values = [] #Alter
x_values = [] #Geschlecht
y_values = [] #Notiz



# Iterate over rows
for row in range(2, max_row):
    # Create an array for the current row
    current_row = []

    # Iterate over columns from B to Y
    for col in range(1, 26):  # 26 corresponds to the column index for Y
        cell_value = ws.cell(row=row, column=col).value
        current_row.append(cell_value)

    # Append the current row to the list of all rows
    all_rows.append(current_row)
    a_values.append(current_row[0])
    b_values.append(current_row[1])
    c_values.append(current_row[2])
    d_values.append(current_row[3])
    e_values.append(current_row[4])
    f_values.append(current_row[5])
    g_values.append(current_row[6])
    h_values.append(current_row[7])
    i_values.append(current_row[8])
    j_values.append(current_row[9])
    k_values.append(current_row[10])
    l_values.append(current_row[11])
    m_values.append(current_row[12])
    n_values.append(current_row[13])
    o_values.append(current_row[14])
    p_values.append(current_row[15])
    q_values.append(current_row[16])
    r_values.append(current_row[17])
    s_values.append(current_row[18])
    t_values.append(current_row[19])
    u_values.append(current_row[20])
    v_values.append(current_row[21])
    w_values.append(current_row[22])
    x_values.append(current_row[23])
    y_values.append(current_row[24])


# Print or use the array of rows as needed
for row in all_rows:
    print(row)

print(a_values)
print(b_values)
print(c_values)
print(d_values)
print(e_values)
print(f_values)
print(g_values)
print(h_values)
print(i_values)
print(j_values)
print(k_values)
print(l_values)
print(m_values)
print(n_values)
print(o_values)
print(p_values)
print(q_values)
print(r_values)
print(s_values)
print(t_values)
print(u_values)
print(v_values)
print(w_values)
print(x_values)
print(y_values)

#statistics
print(Color.RED + "-------------- STATISTICS -----------------------" + Color.RESET)
total_count = len(x_values)
# Count the occurrences of 'M' and 'W'
count_men = x_values.count('M')
count_women = x_values.count('W')
percentage_men = (count_men / total_count) * 100
percentage_women = (count_women / total_count) * 100

# Print the percentages
print(f"Percentage of Men: {percentage_men:.2f}%")
print(f"Percentage of Women: {percentage_women:.2f}%")

